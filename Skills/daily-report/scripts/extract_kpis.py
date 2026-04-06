#!/usr/bin/env python3
"""
BEAM — extract_kpis.py
从数据文件中提取核心 KPI，输出 .tmp/kpis.json。

这是一个模板脚本。每个项目的数据 schema 不同，需要定制以下部分：
  1. DATA_FILE — 数据文件路径
  2. parse_* 函数 — 字段名映射和清洗逻辑
  3. compute_kpis() — KPI 计算公式

当前实现：美妆电商数据看板（订单表 + 日报表 + 评论表）

支持格式：.xlsx / .xls / .csv / .tsv
  - Excel：多工作表，按表名关键词自动识别
  - CSV/TSV：单文件，用文件名作为"工作表名"（如 orders.csv → sheet "orders"）
    若文件名含 "订单"/"order"/"日报"/"daily"/"评论"/"review" 关键词则自动匹配

用法：
  python extract_kpis.py <数据文件路径> [--output .tmp/kpis.json]
"""

import argparse
import json
import os
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

# ── 依赖检查 ──────────────────────────────────────────────────────────
_missing = []
try:
    import openpyxl
except ImportError:
    _missing.append("openpyxl")

try:
    import pandas as pd
    import chardet
except ImportError as e:
    _missing.append("pandas" if "pandas" in str(e) else "chardet")

if _missing:
    print(f"ERROR: 缺少依赖: {', '.join(_missing)}。运行: pip install {' '.join(_missing)}", file=sys.stderr)
    sys.exit(1)


# ── CSV/TSV 适配器（模拟 openpyxl 接口）─────────────────────────────

class _CsvSheet:
    """模拟 openpyxl Worksheet，包装 pandas DataFrame。"""
    def __init__(self, df: "pd.DataFrame"):
        self._df = df

    def iter_rows(self, values_only=True):
        yield tuple(self._df.columns)
        for _, row in self._df.iterrows():
            yield tuple(row)


class _CsvWorkbook:
    """模拟 openpyxl Workbook，供 compute_kpis() 无感使用。"""
    def __init__(self, df: "pd.DataFrame", filepath: str):
        stem = Path(filepath).stem          # 文件名去掉扩展名作为"工作表名"
        self._sheets = {stem: _CsvSheet(df)}
        self.sheetnames = [stem]

    def __getitem__(self, name):
        return self._sheets.get(name)


def _detect_encoding(filepath: str) -> str:
    with open(filepath, "rb") as f:
        raw = f.read(65536)
    result = chardet.detect(raw)
    return result.get("encoding") or "utf-8"


# ── 数据读取（统一入口）──────────────────────────────────────────────

def load_data(filepath: str):
    """
    统一加载数据文件，返回 workbook-like 对象（支持 xlsx/xls/csv/tsv）。
    Excel 返回真实 openpyxl Workbook；CSV/TSV 返回 _CsvWorkbook 适配器。
    """
    if not os.path.exists(filepath):
        print(f"ERROR: 数据文件不存在: {filepath}", file=sys.stderr)
        sys.exit(1)

    ext = Path(filepath).suffix.lower()

    if ext in (".xlsx", ".xls"):
        return openpyxl.load_workbook(filepath, data_only=True)

    if ext in (".csv", ".tsv"):
        sep = "\t" if ext == ".tsv" else ","
        for enc in ["utf-8", None, "latin-1"]:
            try:
                if enc is None:
                    enc = _detect_encoding(filepath)
                df = pd.read_csv(filepath, sep=sep, encoding=enc, low_memory=False)
                break
            except (UnicodeDecodeError, LookupError):
                continue
        else:
            print(f"ERROR: 无法解码文件 {filepath}", file=sys.stderr)
            sys.exit(1)
        return _CsvWorkbook(df, filepath)

    print(f"ERROR: 不支持的文件格式 {ext}。支持: .xlsx .xls .csv .tsv", file=sys.stderr)
    sys.exit(1)



def parse_num(s):
    """清洗数值字符串：去除 ¥、逗号、元、空格、%"""
    if s is None:
        return 0.0
    s = str(s)
    s = re.sub(r'[¥,元\s%]', '', s)
    try:
        return float(s)
    except ValueError:
        return 0.0


def norm_channel(raw):
    """渠道名标准化（与看板 JS 端 normCh 保持一致）"""
    if not raw:
        return "其他"
    raw = str(raw).strip()
    mapping = {
        "淘宝": ["淘宝", "taobao", "TB"],
        "天猫": ["天猫", "tmall", "TM"],
        "京东": ["京东", "jd", "JD", "jingdong"],
        "抖音": ["抖音", "douyin", "DY", "抖音商城"],
        "拼多多": ["拼多多", "pdd", "PDD", "pinduoduo"],
        "小红书": ["小红书", "RED", "red", "xiaohongshu"],
        "快手": ["快手", "kuaishou", "KS"],
        "微信": ["微信", "wechat", "wx", "微信小店"],
        "线下": ["线下", "offline", "门店"],
    }
    for standard, variants in mapping.items():
        if raw in variants or raw.lower() in [v.lower() for v in variants]:
            return standard
    return raw


def sheet_to_dicts(ws):
    """将工作表转为字典列表（首行为表头）。兼容 openpyxl Worksheet 和 _CsvSheet。"""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:] if any(c is not None for c in row)]


# ── KPI 计算（项目定制区）──────────────────────────────────────────────
# >>> 以下函数按具体项目的数据 schema 定制 <<<

def compute_kpis(wb):
    """
    从 workbook 提取 KPI。

    期望的工作表：
      - 订单明细（含：下单日期, 渠道, 实收金额, 毛利, 退款标记, 商品名称）
      - 日报汇总（含：日期, 总GMV(元), 毛利率(%), 总订单数, 广告ROI, 新客数, 复购客数, 平均评分）
      - 评论数据（含：星级评分）

    如果你的数据结构不同，修改此函数中的字段名映射即可。
    """
    result = {}

    # 自动检测工作表（按名称关键词匹配）
    sheet_names = wb.sheetnames
    orders_ws = daily_ws = reviews_ws = None

    for name in sheet_names:
        lower = name.lower()
        if "订单" in name or "order" in lower:
            orders_ws = wb[name]
        elif "日报" in name or "daily" in lower or "汇总" in name:
            daily_ws = wb[name]
        elif "评论" in name or "review" in lower or "评价" in name:
            reviews_ws = wb[name]

    # ── 概览 KPI（来自订单表）──
    if orders_ws:
        orders = sheet_to_dicts(orders_ws)
        total_revenue = sum(parse_num(o.get("实收金额", 0)) for o in orders)
        total_cost = sum(parse_num(o.get("毛利", 0)) for o in orders)
        order_count = len(orders)
        refund_count = sum(1 for o in orders if str(o.get("退款标记", "")).strip() in ("是", "已退款", "1", "true", "True"))
        avg_order_value = total_revenue / order_count if order_count > 0 else 0
        refund_rate = (refund_count / order_count * 100) if order_count > 0 else 0
        gross_margin = (total_cost / total_revenue * 100) if total_revenue > 0 else 0

        result["overview"] = {
            "total_gmv": round(total_revenue, 2),
            "gross_margin_pct": round(gross_margin, 1),
            "avg_order_value": round(avg_order_value, 2),
            "refund_rate_pct": round(refund_rate, 1),
            "order_count": order_count,
        }

        # ── 渠道分布（TOP5）──
        channel_gmv = {}
        for o in orders:
            ch = norm_channel(o.get("渠道"))
            channel_gmv[ch] = channel_gmv.get(ch, 0) + parse_num(o.get("实收金额", 0))
        top_channels = sorted(channel_gmv.items(), key=lambda x: -x[1])[:5]
        result["channels"] = [{"name": ch, "gmv": round(v, 2)} for ch, v in top_channels]

        # ── TOP3 热销商品 ──
        sku_gmv = {}
        for o in orders:
            sku = str(o.get("商品名称", "未知")).strip()
            sku_gmv[sku] = sku_gmv.get(sku, 0) + parse_num(o.get("实收金额", 0))
        top_skus = sorted(sku_gmv.items(), key=lambda x: -x[1])[:3]
        result["top3_skus"] = [{"name": s, "gmv": round(v, 2)} for s, v in top_skus]

        # ── 近 7 日环比 ──
        try:
            dated_orders = []
            for o in orders:
                d = o.get("下单日期")
                if isinstance(d, datetime):
                    dated_orders.append((d, parse_num(o.get("实收金额", 0))))
                elif isinstance(d, str):
                    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
                        try:
                            dated_orders.append((datetime.strptime(d.strip(), fmt), parse_num(o.get("实收金额", 0))))
                            break
                        except ValueError:
                            continue

            if dated_orders:
                max_date = max(d for d, _ in dated_orders)
                last_7d = sum(v for d, v in dated_orders if d > max_date - timedelta(days=7))
                prev_7d = sum(v for d, v in dated_orders if max_date - timedelta(days=14) < d <= max_date - timedelta(days=7))
                wow_pct = ((last_7d - prev_7d) / prev_7d * 100) if prev_7d > 0 else 0
                result["last_7d"] = {
                    "gmv": round(last_7d, 2),
                    "wow_pct": round(wow_pct, 1),
                }
        except Exception:
            pass

    # ── 今日实时（来自日报表最新行）──
    if daily_ws:
        daily = sheet_to_dicts(daily_ws)
        if daily:
            latest = daily[-1]
            result["today"] = {
                "gmv": parse_num(latest.get("总GMV(元)", 0)),
                "orders": int(parse_num(latest.get("总订单数", 0))),
                "ad_roi": round(parse_num(latest.get("广告ROI", 0)), 2),
                "gross_margin_pct": round(parse_num(latest.get("毛利率(%)", 0)), 1),
                "new_customers": int(parse_num(latest.get("新客数", 0))),
                "repeat_customers": int(parse_num(latest.get("复购客数", 0))),
            }
            # 日期区间
            dates = []
            for row in daily:
                d = row.get("日期")
                if isinstance(d, datetime):
                    dates.append(d)
            if dates:
                result["date_range"] = {
                    "start": min(dates).strftime("%Y-%m-%d"),
                    "end": max(dates).strftime("%Y-%m-%d"),
                }

    # ── 评论评分 ──
    if reviews_ws:
        reviews = sheet_to_dicts(reviews_ws)
        scores = [parse_num(r.get("星级评分", 0)) for r in reviews if parse_num(r.get("星级评分", 0)) > 0]
        if scores:
            result["reviews"] = {
                "avg_score": round(sum(scores) / len(scores), 1),
                "count": len(scores),
            }

    # ── 元信息 ──
    result["meta"] = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "data_file": os.path.basename(args.data_file) if 'args' in globals() else "unknown",
    }

    return result


# ── 主入口 ──────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="BEAM: 从数据文件提取核心 KPI")
    parser.add_argument("data_file", help="Excel 数据文件路径")
    parser.add_argument("--output", default=".tmp/kpis.json", help="输出 JSON 路径（默认 .tmp/kpis.json）")

    global args
    args = parser.parse_args()

    # 确保输出目录存在
    os.makedirs(os.path.dirname(args.output) or ".", exist_ok=True)

    wb = load_data(args.data_file)
    kpis = compute_kpis(wb)

    # 更新元信息
    kpis["meta"]["data_file"] = os.path.basename(args.data_file)

    with open(args.output, "w", encoding="utf-8") as f:
        json.dump(kpis, f, ensure_ascii=False, indent=2)

    print(f"KPI 提取完成 → {args.output}")
    print(f"  概览: GMV=¥{kpis.get('overview', {}).get('total_gmv', 'N/A')}")


if __name__ == "__main__":
    main()
